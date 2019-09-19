# -*- coding: utf-8 -*-
# @Time    : 2018/5/6 9:45
# @Author  : chenjinlin
# @Email   : chenjinlin@163.com
# @File    : cjl_excel.py
# @Software: PyCharm
import os
import json
import openpyxl
import xlsxwriter
from utils import reuse_methods
from openpyxl import load_workbook
import constants

# def xlsreader(filename):
#     workbook = load_workbook(filename=filename)
#     sheets = workbook.get_sheet_names()
#     sheet = workbook.get_sheet_by_name(sheets[0])
#     return sheet

class RW_EXCEL:
    # 打开workbook
    def openworkbook(self,filename):
        workbook = load_workbook(filename)
        # print(workbook.sheetnames)
        return workbook

    # 打开页签
    def opnensheet(self,filename, sheetname):
        wb = RW_EXCEL.openworkbook(self,filename)
        sheet = wb[sheetname]
        # print (sheet)
        return sheet

    # 获取每行数据，返回字典
    def getRowValue(self,filename, sheetname, n):
        sheet = RW_EXCEL.opnensheet(self,filename, sheetname)
        row_dict = {}
        titlename = [sheet["A1"].value, sheet["B1"].value, sheet["C1"].value, sheet["D1"].value, sheet["E1"].value,
                     sheet["F1"].value]
        # print (titlename)
        i = -1
        for rownvalue in sheet[n]:
            # print(rownvalue.value)
            i += 1
            row_dict[titlename[i]] = rownvalue.value
        return row_dict




if __name__ == '__main__':
    filename =constants.CONSTANTS.FILE_NAME
    sheetname='beehiveApiTestCase'
    n='9'
    # wb = openworkbook(filename)
    # sheet = opnensheet(filename,sheetname)
    rw=RW_EXCEL()
    rv=rw.getRowValue(filename,sheetname,n)
    print (rv)


    # print (sheet["A2"].value)
    #
    # print (sheet.max_row)
    # print (sheet.max_column)

